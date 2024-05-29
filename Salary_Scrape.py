import selenium
from selenium import webdriver
import pandas as pd
import xlrd
from selenium.webdriver.common.keys import Keys
#from selenium.webdriver.common.by import By
#from selenium.webdriver.remote.webelement import WebElement
#from selenium.webdriver.support.ui import WebDriverWait
#from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import csv
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException

#read rows from excel
path = "C:/Users/poorn/OneDrive - dadoshealth.com/Centrak/Engage/ROI_Value Projection/OneDrive_2022-06-15/Projected Value/Temp_Sal.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
rows = sheet.max_row
columns = sheet.max_column
lst =[]
for i in range(2,rows+1):
    for j in range(1,columns+1):
        z=(sheet.cell(row=i,column=j).value)
        lst.append(z)
lst2=[]
for i in range(2,rows+1):
    for j in range(2,columns+1):
        z=(sheet.cell(row=i,column=j).value)
        lst2.append(z)
#driver path
DRIVER_PATH = 'C:/Users/poorn/OneDrive/Desktop/chromedriver_win32/chromedriver.exe'
driver = webdriver.Chrome(executable_path=DRIVER_PATH)
#open page
driver.get('https://www.salary.com/')
driver.maximize_window()
import time
time.sleep(2)
r=1
table=[]
for i in lst:
    for j in lst2:
        time.sleep(2)
        input_txt_title = driver.find_element_by_xpath('//*[@id="trafficdrivertad-worth-jobtitle_input"]')
        input_txt_location = driver.find_element_by_xpath('//*[@id="trafficdrivertad-worth-location_input"]')
        input_txt_title.send_keys(i)
        input_txt_location.send_keys(j)
        driver.find_element_by_xpath('//*[@id="trafficdrivertad-worth-jobtitle_input"]').send_keys
        driver.find_element_by_xpath('//*[@id="trafficdrivertad-worth-location_input"]').send_keys(Keys.ENTER)
        try:
            driver.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[5]/div[4]/div[1]/a').click()
            time.sleep(10)
            driver.find_element('//*[@id="linkchart"]').click()
            time.sleep(4)
            driver.find_element_by_xpath('//*[@id="sal-demoform-popup"]').click()
            time.sleep(4)
            med_sal = driver.find_element_by_xpath('//*[@id="divtable"]/div[2]/table/tbody/tr[3]/td[2]')
            driver.find_element_by_xpath('//*[@id="payinterval"]').click()
            driver.find_element_by_xpath('//*[@id="payinterval"]/option[6]').click()
            med_sal_hourly = driver.find_element_by_xpath('//*[@id="divtable"]/div[2]/table/tbody/tr[3]/td[2]')
            role = driver.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[2]/div/div[3]/div/h1/b')
            Table_data = {
                'Role': role.text,
                'Annual Pay': med_sal.text,
                'Hourly Pay': med_sal_hourly.text}
            table.append(Table_data)
            df = pd.DataFrame(table)
            r += 1
            driver.get('https://www.salary.com/')
            time.sleep(5)
        except Exception as ex:
            pass

# write to excel file
df.to_csv('Salary_Scrapped.csv')
driver.close()


